import contextlib
import os
import sqlite3
import tempfile
import unittest

from openpyxl import Workbook

from graph_repository import (
    COMMAND_COLUMNS,
    END_NODE_ID,
    START_NODE_ID,
    GraphRepository,
    GraphSchemaError,
    GraphValidationError,
    WorkbookValidationError,
)
from instructions.models import CommandRecord, InstructionDraft
from instructions.registry import INSTRUCTION_SPECS
from 数据库操作 import DatabaseOperation


VALID_TYPES = {"图像点击", "时间等待", "文本输入"}


class GraphRepositoryTests(unittest.TestCase):
    def setUp(self):
        self.temporary_directory = tempfile.TemporaryDirectory()
        self.database_path = os.path.join(
            self.temporary_directory.name, "commands.db"
        )
        self.database = DatabaseOperation(self.database_path)
        self.repository = GraphRepository(
            self.database_path, valid_type_ids=VALID_TYPES
        )

    def tearDown(self):
        self.temporary_directory.cleanup()

    @staticmethod
    def draft(type_id="图像点击", **parameters):
        return InstructionDraft(
            type_id=type_id,
            parameters=parameters,
            repeat_count=2,
            error_policy="提示异常并暂停",
            note="测试",
        )

    def test_fresh_schema_contains_fixed_boundary_chain(self):
        with contextlib.closing(sqlite3.connect(self.database_path)) as connection:
            command_columns = tuple(
                row[1] for row in connection.execute("PRAGMA table_info('命令')")
            )
            node_columns = tuple(
                row[1] for row in connection.execute("PRAGMA table_info('节点')")
            )
            edge_columns = tuple(
                row[1]
                for row in connection.execute("PRAGMA table_info('节点连接')")
            )
            foreign_key_errors = connection.execute(
                "PRAGMA foreign_key_check"
            ).fetchall()
        self.assertEqual(command_columns, COMMAND_COLUMNS)
        self.assertEqual(node_columns, ("节点ID", "命令ID", "节点类型", "X", "Y"))
        self.assertEqual(edge_columns, ("源节点ID", "目标节点ID"))
        self.assertEqual(foreign_key_errors, [])

        snapshot = self.repository.snapshot()
        self.assertEqual(snapshot.commands, ())
        self.assertEqual(
            [node.node_id for node in snapshot.nodes],
            [START_NODE_ID, END_NODE_ID],
        )
        self.assertEqual(
            [(edge.source, edge.target) for edge in snapshot.edges],
            [(START_NODE_ID, END_NODE_ID)],
        )

    def test_unknown_legacy_command_schema_is_rejected_without_data_loss(self):
        legacy_path = os.path.join(self.temporary_directory.name, "legacy.db")
        with contextlib.closing(sqlite3.connect(legacy_path)) as connection:
            connection.execute(
                "CREATE TABLE 命令("
                "ID INTEGER PRIMARY KEY, 图像名称 TEXT, 指令类型 TEXT, "
                "参数1 TEXT, 参数2 TEXT, 参数3 TEXT, 参数4 TEXT, "
                "重复次数 INTEGER, 异常处理 TEXT, 备注 TEXT)"
            )
            connection.execute(
                "INSERT INTO 命令 VALUES "
                "(1, NULL, '图像点击', NULL, NULL, NULL, NULL, 1, '', '')"
            )
            connection.commit()

        with self.assertRaises(GraphSchemaError):
            DatabaseOperation(legacy_path)
        with contextlib.closing(sqlite3.connect(legacy_path)) as connection:
            self.assertEqual(
                connection.execute("SELECT ID, 指令类型 FROM 命令").fetchall(),
                [(1, "图像点击")],
            )
            self.assertIsNone(
                connection.execute(
                    "SELECT 1 FROM sqlite_master WHERE name='节点'"
                ).fetchone()
            )

    def test_add_update_duplicate_delete_and_clear_keep_one_chain(self):
        first = self.repository.add_command(self.draft(path="一.png"), x=100, y=40)
        second = self.repository.add_command(
            {
                "type_id": "时间等待",
                "parameters": {"seconds": 1.5},
                "repeat_count": 1,
                "error_policy": "自动跳过",
                "note": "等待",
            },
            x=200,
            y=40,
        )
        self.assertIsInstance(first, CommandRecord)
        self.assertEqual([item.id for item in self.repository.list_commands()], [first.id, second.id])
        self.assertEqual([item.order for item in self.repository.list_commands()], [0, 1])

        updated = self.repository.update_command(
            first.id,
            self.draft(type_id="文本输入", text="你好"),
        )
        self.assertEqual(updated.id, first.id)
        self.assertEqual(updated.type_id, "文本输入")
        self.assertEqual(updated.parameters, {"text": "你好"})

        duplicate = self.repository.duplicate_command(first.id)
        self.assertNotEqual(duplicate.id, first.id)
        self.assertEqual(
            [item.id for item in self.repository.list_commands()],
            [first.id, duplicate.id, second.id],
        )
        self.assertEqual(self.repository.delete_commands([first.id, 999999]), 1)
        self.assertEqual(
            [item.id for item in self.repository.list_commands()],
            [duplicate.id, second.id],
        )
        self.repository.validate_graph()

        self.repository.clear()
        snapshot = self.repository.snapshot()
        self.assertEqual(snapshot.commands, ())
        self.assertEqual(
            [(edge.source, edge.target) for edge in snapshot.edges],
            [(START_NODE_ID, END_NODE_ID)],
        )

    def test_split_edge_reorder_and_position_persistence(self):
        first = self.repository.add_command(self.draft(path="一.png"))
        first_node = next(
            node
            for node in self.repository.snapshot().nodes
            if node.command_id == first.id
        )
        inserted = self.repository.add_command(
            self.draft(type_id="时间等待", seconds=2),
            split_edge=(START_NODE_ID, first_node.node_id),
            x=80,
            y=90,
        )
        self.assertEqual(
            [item.id for item in self.repository.list_commands()],
            [inserted.id, first.id],
        )

        self.repository.reorder_chain([first.id, inserted.id])
        self.assertEqual(
            [item.id for item in self.repository.list_commands()],
            [first.id, inserted.id],
        )
        self.repository.save_node_positions(
            {START_NODE_ID: (-20, 15), first_node.node_id: (123.5, 456.25)}
        )
        positions = {
            node.node_id: (node.x, node.y)
            for node in self.repository.snapshot().nodes
        }
        self.assertEqual(positions[START_NODE_ID], (-20.0, 15.0))
        self.assertEqual(positions[first_node.node_id], (123.5, 456.25))

        with self.assertRaises(GraphValidationError):
            self.repository.reorder_chain([first.id])
        self.assertEqual(
            [item.id for item in self.repository.list_commands()],
            [first.id, inserted.id],
        )

    def test_reorder_and_position_commit_rolls_back_as_one_transaction(self):
        first = self.repository.add_command(self.draft(path="一.png"))
        second = self.repository.add_command(self.draft(path="二.png"))
        original_ids = [item.id for item in self.repository.list_commands()]

        with self.assertRaises(KeyError):
            self.repository.reorder_chain_and_save_positions(
                [second.id, first.id],
                {"missing-node": (1.0, 2.0)},
            )

        self.assertEqual(
            [item.id for item in self.repository.list_commands()], original_ids
        )
        self.repository.validate_graph()

    def test_snapshot_loads_directly_into_node_editor_widget(self):
        os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
        from PySide6.QtWidgets import QApplication
        from node_editor.widget import NodeEditorWidget

        application = QApplication.instance() or QApplication([])
        command = self.repository.add_command(self.draft())
        snapshot = self.repository.snapshot()
        widget = NodeEditorWidget()
        widget.load_graph(snapshot.nodes, snapshot.edges, INSTRUCTION_SPECS)
        self.assertEqual(
            [
                node.command_id
                for node in widget.scene.chain_order
                if not node.is_terminal
            ],
            [command.id],
        )
        terminals = {
            node.terminal_role
            for node in widget.scene.nodes_by_id.values()
            if node.terminal_role is not None
        }
        self.assertEqual(terminals, {"start", "end"})
        widget.deleteLater()
        application.processEvents()

    def test_database_tampering_is_detected(self):
        command = self.repository.add_command(self.draft())
        with contextlib.closing(sqlite3.connect(self.database_path)) as connection:
            connection.execute("PRAGMA foreign_keys=ON")
            node_id = connection.execute(
                "SELECT 节点ID FROM 节点 WHERE 命令ID=?", (command.id,)
            ).fetchone()[0]
            connection.execute("DELETE FROM 节点连接 WHERE 源节点ID=?", (node_id,))
            connection.commit()
        with self.assertRaisesRegex(GraphValidationError, "完整连线|各有一个"):
            self.repository.validate_graph()

    def test_workbook_round_trip_replaces_graph_and_settings(self):
        first = self.repository.add_command(self.draft(path="图.png"), x=100, y=50)
        second = self.repository.add_command(
            self.draft(type_id="时间等待", seconds=3), x=200, y=50
        )
        self.database.set_setting_value("测试设置", "来源")
        workbook = Workbook()
        self.repository.export_to_workbook(workbook, self.database)
        self.assertEqual(workbook.sheetnames, ["命令", "节点", "连线", "设置"])

        target_path = os.path.join(self.temporary_directory.name, "target.db")
        target_database = DatabaseOperation(target_path)
        target_repository = GraphRepository(
            target_path, valid_type_ids=VALID_TYPES
        )
        target_repository.import_from_workbook(workbook)
        self.assertEqual(
            [item.id for item in target_repository.list_commands()],
            [first.id, second.id],
        )
        self.assertEqual(
            [item.parameters for item in target_repository.list_commands()],
            [{"path": "图.png"}, {"seconds": 3}],
        )
        self.assertEqual(target_database.get_setting_value("测试设置"), "来源")
        target_repository.validate_graph()

    def test_invalid_workbooks_are_rejected_atomically(self):
        original = self.repository.add_command(self.draft(path="保留.png"))
        self.database.set_setting_value("测试设置", "保留")
        workbook = Workbook()
        self.repository.export_to_workbook(workbook, self.database)
        workbook["命令"].cell(2, 3).value = "{'not': 'json'}"

        with self.assertRaises(WorkbookValidationError):
            self.repository.import_from_workbook(workbook)
        self.assertEqual(
            [item.id for item in self.repository.list_commands()], [original.id]
        )
        self.assertEqual(self.database.get_setting_value("测试设置"), "保留")

        old_workbook = Workbook()
        old_sheet = old_workbook.active
        old_sheet.title = "命令"
        old_sheet.append(
            [
                "ID", "图像名称", "指令类型", "参数信息", "参数-2",
                "参数-3", "参数-4", "重复次数", "异常处理", "备注",
            ]
        )
        with self.assertRaises(WorkbookValidationError):
            self.repository.import_from_workbook(old_workbook)
        self.assertEqual(
            [item.id for item in self.repository.list_commands()], [original.id]
        )

    def test_illegal_graph_in_workbook_does_not_change_database(self):
        original = self.repository.add_command(self.draft())
        workbook = Workbook()
        self.repository.export_to_workbook(workbook, self.database)
        edge_sheet = workbook["连线"]
        edge_sheet.delete_rows(2, edge_sheet.max_row)
        edge_sheet.append((START_NODE_ID, END_NODE_ID))

        with self.assertRaises(WorkbookValidationError):
            self.repository.import_from_workbook(workbook)
        self.assertEqual(
            [item.id for item in self.repository.list_commands()], [original.id]
        )
        self.repository.validate_graph()


if __name__ == "__main__":
    unittest.main()
