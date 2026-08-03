import contextlib
import os
import sqlite3
import sys
import tempfile
import unittest
from unittest.mock import patch

from openpyxl import Workbook

import functions
from 数据库操作 import DatabaseOperation, REMOVED_COMMAND_TYPES


class DataPathTests(unittest.TestCase):
    def test_source_data_is_under_project_folder(self):
        self.assertEqual(functions.DATA_FOLDER, os.path.join(functions.INSTALL_FOLDER, "data"))
        self.assertEqual(functions.DATABASE_PATH, os.path.join(functions.DATA_FOLDER, "命令集.db"))

    def test_frozen_install_and_resource_folders_are_separate(self):
        executable = os.path.join("D:\\Portable Clicker", "Clicker.exe")
        resource_folder = os.path.join("C:\\Temp", "_MEI123")
        with patch.object(sys, "frozen", True, create=True), \
                patch.object(sys, "executable", executable), \
                patch.object(sys, "_MEIPASS", resource_folder, create=True):
            self.assertEqual(functions.get_install_folder(), os.path.dirname(executable))
            self.assertEqual(functions.get_resource_folder(), resource_folder)

    def test_portable_resource_path_resolves_from_data(self):
        expected = os.path.normpath(os.path.join(functions.DATA_FOLDER, "images"))
        self.assertEqual(DatabaseOperation.resolve_resource_path("images"), expected)
        self.assertEqual(DatabaseOperation.portable_resource_path(expected), "images")

    def test_database_schema_and_excel_settings_round_trip(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            source = DatabaseOperation(os.path.join(temporary_directory, "source.db"))
            source.set_setting_value("测试值", "已写入")
            source.writes_to_branch_info("测试分支", "K", 3)
            workbook = Workbook()
            source.export_settings_to_excel(workbook)

            target = DatabaseOperation(os.path.join(temporary_directory, "target.db"))
            self.assertTrue(target.import_settings_from_excel(workbook))
            self.assertEqual(target.get_setting_value("测试值"), "已写入")
            self.assertIn(("测试分支", "K", 3), target.get_branch_info())

            with contextlib.closing(sqlite3.connect(target.db_path)) as connection:
                tables = {
                    row[0]
                    for row in connection.execute(
                        "SELECT name FROM sqlite_master WHERE type='table'"
                    )
                }
                setting_columns = [
                    (row[1], row[3], row[5])
                    for row in connection.execute("PRAGMA table_info('设置')")
                ]
                setting_types = dict(
                    connection.execute("SELECT 设置项, 类型 FROM 设置")
                )
            self.assertTrue({"设置", "窗口大小", "分支", "资源文件夹"} <= tables)
            self.assertEqual(
                setting_columns,
                [("类型", 1, 0), ("设置项", 1, 1), ("值", 1, 0)],
            )
            self.assertEqual(setting_types["测试值"], "基础设置")
            self.assertEqual(setting_types["apiKey"], "三方接口")
            self.assertEqual(setting_types["快捷键-开始运行"], "全局快捷键")

            setting_sheet = workbook["设置"]
            exported_settings = {
                row[1]: row[3]
                for row in setting_sheet.iter_rows(min_row=2, values_only=True)
                if row[0] == "设置"
            }
            self.assertEqual(exported_settings["测试值"], "基础设置")
            self.assertEqual(exported_settings["apiKey"], "三方接口")

    def test_removed_commands_are_deleted_from_existing_database(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            database_path = os.path.join(temporary_directory, "commands.db")
            with contextlib.closing(sqlite3.connect(database_path)) as connection:
                connection.execute("CREATE TABLE 命令(指令类型 TEXT NOT NULL)")
                connection.executemany(
                    "INSERT INTO 命令 VALUES (?)",
                    [(command_type,) for command_type in REMOVED_COMMAND_TYPES]
                    + [("图像点击",)],
                )
                connection.commit()

            DatabaseOperation(database_path)

            with contextlib.closing(sqlite3.connect(database_path)) as connection:
                command_types = connection.execute(
                    "SELECT 指令类型 FROM 命令 ORDER BY 指令类型"
                ).fetchall()
            self.assertEqual(command_types, [("图像点击",)])

    def test_legacy_excel_settings_are_not_imported(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "设置"
        sheet.append(["[Config]", None])
        sheet.append(["测试值", "不应写入"])
        with tempfile.TemporaryDirectory() as temporary_directory:
            database = DatabaseOperation(os.path.join(temporary_directory, "test.db"))
            self.assertFalse(database.import_settings_from_excel(workbook))
            self.assertIsNone(database.get_setting_value("测试值"))

    def test_excel_settings_without_category_are_classified_on_import(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "设置"
        sheet.append(["类型", "名称", "值", "附加值", "排序"])
        sheet.append(["设置", "apiKey", "旧格式接口值", None, None])
        sheet.append(["设置", "快捷键-开始运行", "ctrl+f10", None, None])
        with tempfile.TemporaryDirectory() as temporary_directory:
            database = DatabaseOperation(os.path.join(temporary_directory, "test.db"))
            self.assertTrue(database.import_settings_from_excel(workbook))
            with contextlib.closing(sqlite3.connect(database.db_path)) as connection:
                setting_types = dict(
                    connection.execute("SELECT 设置项, 类型 FROM 设置")
                )
            self.assertEqual(setting_types["apiKey"], "三方接口")
            self.assertEqual(setting_types["快捷键-开始运行"], "全局快捷键")

    def test_unknown_settings_schema_is_rejected_without_data_loss(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            database_path = os.path.join(temporary_directory, "unknown.db")
            with contextlib.closing(sqlite3.connect(database_path)) as connection:
                connection.execute("CREATE TABLE 设置(未知字段 TEXT)")
                connection.execute("INSERT INTO 设置 VALUES ('保留数据')")
                connection.commit()

            with self.assertRaisesRegex(RuntimeError, "未知的设置表结构"):
                DatabaseOperation(database_path)

            with contextlib.closing(sqlite3.connect(database_path)) as connection:
                self.assertEqual(
                    connection.execute("SELECT 未知字段 FROM 设置").fetchall(),
                    [("保留数据",)],
                )

    def test_legacy_database_global_parameters_are_migrated(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            database_path = os.path.join(temporary_directory, "legacy.db")
            with contextlib.closing(sqlite3.connect(database_path)) as connection:
                connection.execute(
                    "CREATE TABLE 全局参数(资源文件夹路径 TEXT, 分支表名 TEXT)"
                )
                connection.execute(
                    "CREATE TABLE 设置(设置类型 TEXT PRIMARY KEY, 值 TEXT NOT NULL)"
                )
                connection.executemany(
                    "INSERT INTO 设置 VALUES (?, ?)",
                    [
                        ("Clicker", "(640, 480)"),
                        ("激活状态", "True"),
                        ("apiKey", "旧接口值"),
                        ("快捷键-开始运行", "ctrl+f10"),
                    ],
                )
                connection.executemany(
                    "INSERT INTO 全局参数 VALUES (?, ?)",
                    [("images", None), (None, "旧分支")],
                )
                connection.commit()
            database = DatabaseOperation(database_path)
            self.assertIn("旧分支", database.get_branch_info(True))
            self.assertEqual(
                database.extract_resource_folder_path()[0], functions.IMAGES_FOLDER
            )
            self.assertIsNone(database.get_setting_value("Clicker"))
            self.assertEqual(database.get_window_state("Clicker")["size"], (640, 480))
            with contextlib.closing(sqlite3.connect(database_path)) as connection:
                self.assertIsNone(
                    connection.execute(
                        "SELECT 1 FROM sqlite_master WHERE name='全局参数'"
                    ).fetchone()
                )
                setting_rows = {
                    item: (setting_type, value)
                    for setting_type, item, value in connection.execute(
                        "SELECT 类型, 设置项, 值 FROM 设置"
                    )
                }
                setting_columns = [
                    row[1]
                    for row in connection.execute("PRAGMA table_info('设置')")
                ]
            self.assertEqual(setting_columns, ["类型", "设置项", "值"])
            self.assertEqual(setting_rows["激活状态"], ("激活信息", "True"))
            self.assertEqual(setting_rows["apiKey"], ("三方接口", "旧接口值"))
            self.assertEqual(
                setting_rows["快捷键-开始运行"],
                ("全局快捷键", "ctrl+f10"),
            )


if __name__ == "__main__":
    unittest.main()
