import configparser
import ctypes
import os
import sqlite3
import time

import win32con
import win32gui
import winsound
from openpyxl.workbook import Workbook

from functions import get_current_folder


class IniControl:
    def __init__(self):
        self.ini_path = os.path.join(get_current_folder(), "config.ini")
        self.config = configparser.ConfigParser()
        self.config.read(self.ini_path, encoding='utf-8')
        self.db_path = os.path.join(get_current_folder(), "命令集.db")

    @staticmethod
    def get_config() -> configparser.ConfigParser:
        """获取配置文件"""
        config = configparser.ConfigParser()
        config.read("config.ini", encoding="utf-8")
        return config

    def get_setting_data_from_ini(self, selection: str, *args):
        """从ini文件中获取设置数据
        :param selection: 选择的选区域
        :param args: 设置类型参数"""
        try:
            config = self.get_config()
            if not args:
                return None
            elif len(args) == 1:
                return config[selection][args[0]]
            else:
                return {arg: config[selection][arg] for arg in args}
        except Exception as e:
            print(e)
            return {} if len(args) > 1 else None

    def update_settings_in_ini(self, selection: str, **kwargs):
        """更新ini文件中的设置数据
        :param selection: 选择的选区域
        :param kwargs: 设置类型参数"""
        try:
            config = self.get_config()
            for key, value in kwargs.items():
                config[selection][key] = value
            with open("config.ini", "w", encoding="utf-8") as f:
                config.write(f)
        except Exception as e:
            print("更新设置数据失败！", e)

    def get_ocr_info(self) -> dict:
        """从ini中获取百度OCR的API信息"""
        try:
            config = self.get_config()
            return {
                "appId": config["三方接口"]["appid"],
                "apiKey": config["三方接口"]["apikey"],
                "secretKey": config["三方接口"]["secretkey"]
            }
        except Exception as e:
            print("获取OCR信息失败！", e)
            return {}

    @staticmethod
    def get_screen_resolution() -> str:
        """获取屏幕分辨率"""
        user32 = ctypes.windll.user32
        width = user32.GetSystemMetrics(0)
        height = user32.GetSystemMetrics(1)
        return f"{width}*{height}"

    def save_window_size(self, win_width: int, win_height: int, window_name: str):
        """获取窗口大小
        :param win_width: 保存时窗口的宽度
        :param win_height: 保存时窗口的高度
        :param window_name:窗口名称
        :return: 窗口大小"""

        try:
            config = self.get_config()
            save_size = (win_width, win_height)
            window_info = f"{window_name}-{self.get_screen_resolution()}"
            # 检查'窗口大小'选区中是否存在window_name选项
            config["窗口大小"][window_info] = str(save_size)
            with open("config.ini", "w", encoding="utf-8") as f:
                config.write(f)
        except Exception as e:
            print("保存窗口大小失败！", e)

    def set_window_size(self, window):
        def get_window_size(window_name: str):
            """设置窗口大小
            :param window_name:窗口名称
            :return: 窗口大小"""
            window_info = f"{window_name}-{self.get_screen_resolution()}"
            try:
                width_, height_ = eval(self.get_setting_data_from_ini("窗口大小", window_info))
                return int(width_), int(height_)
            except TypeError:
                return 0, 0

        width, height = get_window_size(window.windowTitle())
        if width and height:
            window.resize(width, height)

    def get_global_shortcut(self) -> dict:
        """获取全局快捷键"""
        try:
            config = self.get_config()
            return {
                "开始运行": config["全局快捷键"]["开始运行"].lower().split("+"),
                "结束运行": config["全局快捷键"]["结束运行"].lower().split("+"),
                "分支选择": config["全局快捷键"]["分支选择"].lower().split("+"),
                "暂停和恢复": config["全局快捷键"]["暂停和恢复"].lower().split("+"),
            }
        except Exception as e:
            print("获取全局快捷键失败！", e)
            return {}

    def set_global_shortcut(self, **kwargs):
        """设置全局快捷键
        :param kwargs: 全局快捷键参数, 如：开始运行=["ctrl", "alt", "1"]"""
        try:
            config = self.get_config()
            for key, value in kwargs.items():
                # 将"control"替换为"ctrl"
                value = ['ctrl' if v.lower() == 'control' else v for v in value]
                config["全局快捷键"][key] = "+".join(value).lower()
            with open("config.ini", "w", encoding="utf-8") as f:
                config.write(f)
        except Exception as e:
            print("设置全局快捷键失败！", e)

    def writes_to_resource_folder_path(self, path: str):
        """将资源文件路径写入到config.ini中"""
        try:
            config = self.get_config()
            section = '资源文件夹路径'
            if not config.has_section(section):
                config.add_section(section)
            paths = {key: config.get(section, key) for key in config.options(section)}
            if path not in paths.values():
                keys = sorted(
                    [int(
                        k.replace("路径", "")
                    ) for k in paths.keys() if k.replace("路径", "").isdigit()], reverse=True
                )
                new_key = f"路径{keys[0] + 1 if keys else 1}"
                config.set(section, new_key, path)
                with open("config.ini", "w", encoding="utf-8") as configfile:
                    config.write(configfile)
                print("路径写入成功！")
            else:
                print("路径已经存在！")
        except Exception as e:
            print("写入资源文件路径失败！", e)

    def del_resource_folder_path(self, path: str):
        """删除资源文件路径"""
        try:
            config = self.get_config()
            section = '资源文件夹路径'
            # 检查 '资源文件夹路径' 部分是否存在
            if not config.has_section(section):
                print("配置文件中不存在资源文件夹路径部分！")
                return
            # 获取所有路径键值对并检查路径是否存在
            paths = {key: config.get(section, key) for key in config.options(section)}
            if path not in paths.values():
                print("路径不存在于配置文件中！")
                return
            # 删除指定路径并重新编号
            new_paths = {f"路径{i + 1}": value for i, value in enumerate(v for k, v in paths.items() if v != path)}
            # 清除原有部分并重新添加整理后的路径键值对
            config.remove_section(section)
            config.add_section(section)
            for key, value in new_paths.items():
                config.set(section, key, value)
            # 保存配置文件
            with open("config.ini", "w", encoding="utf-8") as configfile:
                config.write(configfile)
            print("路径删除成功！")
        except Exception as e:
            print("删除资源文件路径失败！", e)

    def move_resource_folder_up_and_down(self, path: str, direction: str):
        """将资源文件夹路径上移或下移
        :param path: 选中的路径
        :param direction: 移动方向（up: 上移, down: 下移）"""
        try:
            config = self.get_config()
            section = '资源文件夹路径'
            if not config.has_section(section):
                print("配置文件中不存在资源文件夹路径部分！")
                return

            paths = {key: config.get(section, key) for key in config.options(section)}
            if path not in paths.values():
                print("路径不存在于配置文件中！")
                return
            path_key = next((key for key, value in paths.items() if value == path), None)
            path_index = list(paths.keys()).index(path_key)
            if direction == 'up' and path_index > 0:
                paths[path_key], paths[
                    list(paths.keys())[path_index - 1]] = paths[list(paths.keys())[path_index - 1]], paths[path_key
                ]
            elif direction == 'down' and path_index < len(paths) - 1:
                paths[path_key], paths[
                    list(paths.keys())[path_index + 1]] = paths[list(paths.keys())[path_index + 1]], paths[path_key
                ]
            config.remove_section(section)
            config.add_section(section)
            for key, value in paths.items():
                config.set(section, key, value)

            with open("config.ini", "w", encoding="utf-8") as configfile:
                config.write(configfile)
            print("路径移动成功！")
        except Exception as e:
            print("移动资源文件夹路径失败！", e)

    def extract_resource_folder_path(self) -> list:
        """提取资源文件夹路径"""
        try:
            config = self.get_config()
            section = '资源文件夹路径'
            if not config.has_section(section):
                return []
            paths = {key: config.get(section, key) for key in config.options(section)}
            return list(paths.values())
        except Exception as e:
            print("提取资源文件夹路径失败！", e)
            return []

    def get_all_png_images_from_resource_folders(self) -> list:
        """Retrieve all PNG image names from resource folder paths defined in the config."""
        try:
            config = self.get_config()
            if '资源文件夹路径' not in config:
                return []
            paths = [config.get('资源文件夹路径', key) for key in config.options('资源文件夹路径')]
            return [file for path in paths for _, _, files in os.walk(path) for file in files if file.endswith('.png')]
        except Exception as e:
            print(f"Failed to retrieve PNG image names: {e}")
            return []

    def matched_complete_path_from_resource_folders(self, file_name: str) -> str:
        """从资源文件夹中匹配完整路径
        :param file_name: 文件名
        :return: 完整路径"""
        try:
            config = self.get_config()
            if '资源文件夹路径' not in config:
                return ''
            paths = [config.get('资源文件夹路径', key) for key in config.options('资源文件夹路径')]
            for path in paths:
                for root, _, files in os.walk(path):
                    for file in files:
                        if file == file_name:
                            return os.path.join(root, file)
            return ''
        except Exception as e:
            print(f"匹配资源文件夹路径失败: {e}")
            return ''

    def writes_to_branch_info(self, branch_name: str, shortcut_key: str, repeat_times: int = 1) -> bool:
        """将分支信息写入到config.ini中
        :param branch_name: 分支名称
        :param shortcut_key: 快捷键
        :param repeat_times: 重复次数
        :return: 如果添加的分支名称已经存在则返回False，添加成功返回True
        """
        try:
            config = self.get_config()
            section = '分支'
            # 如果“分支”部分不存在，则添加该部分
            if not config.has_section(section):
                config.add_section(section)
                config.set(section, "主流程", f"('',{repeat_times})")  # 添加主流程
            # 检查分支名称是否已经存在
            if config.has_option(section, branch_name) and branch_name != "主流程":
                # 如果分支名称存在但快捷键不同，则更新快捷键
                existing_value = eval(config.get(section, branch_name))
                existing_shortcut = existing_value[0]
                existing_repeat_times = existing_value[1]
                if existing_shortcut != shortcut_key:
                    config.set(section, branch_name, f"('{shortcut_key}',{repeat_times})")
                elif existing_repeat_times != repeat_times:
                    config.set(section, branch_name, f"('{shortcut_key}',{repeat_times})")
                else:
                    return False
            else:
                # 将分支名称和快捷键写入到“分支”部分
                config.set(section, branch_name, f"('{shortcut_key}',{repeat_times})")
            # 将更新后的配置写回文件
            with open('config.ini', 'w', encoding='utf-8') as configfile:
                config.write(configfile)
            return True
        except Exception as e:
            print(f"An error occurred: {e}")
            return False

    def set_branch_repeat_times(self, branch_name: str, repeat_times: int) -> None:
        """设置分支的重复次数
        :param branch_name: 分支名称
        :param repeat_times: 重复次数"""
        try:
            config = self.get_config()
            section = '分支'
            if config.has_option(section, branch_name):
                config.set(section, branch_name, f"('{eval(config.get(section, branch_name))[0]}',{repeat_times})")
                with open('config.ini', 'w', encoding='utf-8') as configfile:
                    config.write(configfile)
        except Exception as e:
            print(f"设置分支重复次数失败: {e}")

    def get_branch_repeat_times(self, branch_name: str) -> int:
        """获取分支的重复次数
        :param branch_name: 分支名称
        :return: 分支的重复次数"""
        try:
            config = self.get_config()
            section = '分支'
            if config.has_option(section, branch_name):
                return eval(config.get(section, branch_name))[1]
            return 1
        except Exception as e:
            print(f"获取分支重复次数失败: {e}")
            return 1

    def del_branch_info(self, branch_name: str) -> bool:
        """删除分支信息
        :param branch_name: 分支名称
        :return: 如果分支名称不存在则返回False，删除成功返回True
        """

        def del_branch_in_database():
            """删除数据库中的分支"""
            with sqlite3.connect(self.db_path) as con:
                cursor = con.cursor()
                cursor.execute(
                    "delete from 命令 where 隶属分支=?", (branch_name,)
                )  # 从命令表中删除分支指令
                con.commit()

        try:
            config = self.get_config()
            section = '分支'
            # 如果“分支”部分不存在，则返回False
            if not config.has_section(section):
                return False
            # 检查分支名称是否存在，不存在则返回False
            if not config.has_option(section, branch_name):
                return False
            # 如果为主流程则不允许删除
            if branch_name == "主流程":
                return False
            # 删除指定分支名称
            config.remove_option(section, branch_name)
            del_branch_in_database()  # 删除数据库中的分支
            # 将更新后的配置写回文件
            with open('config.ini', 'w', encoding='utf-8') as configfile:
                config.write(configfile)
            return True
        except Exception as e:
            print(f"删除分支信息失败: {e}")
            return False

    def get_branch_info(self, keys_only: bool = False) -> list:
        """获取分支信息
        :param keys_only: 如果为True，只返回键（分支名称），否则返回键值对（分支名称和快捷键）
        :return: 根据keys_only参数返回全部信息或仅返回键
        """
        try:
            config = self.get_config()
            section = '分支'
            if not config.has_section(section):
                return []
            if keys_only:
                return [key for key, value in config.items(section)]
            else:
                result = []
                for key, value in config.items(section):
                    # 将字符串转换为元组并展开
                    shortcut, priority = eval(value)
                    result.append((key, shortcut, priority))
                return result
        except Exception as e:
            print(f"写入分支信息失败: {e}")
            return []

    def move_branch_info(self, branch_name: str, direction: str) -> bool:
        """移动分支信息
        :param branch_name: 分支名称
        :param direction: 移动方向（up: 上移, down: 下移）
        :return: 移动成功返回 True，否则返回 False
        """
        config = self.get_config()
        if '分支' not in config:
            return False
        branches = list(config['分支'].items())
        index = next((i for i, (name, _) in enumerate(branches) if name == branch_name), None)
        if index is None or branch_name == '主流程':
            return False
        new_index = index - 1 if direction == 'up' and index > 1 else index + 1 if direction == 'down' and index < len(
            branches) - 1 else index
        if new_index != index:
            branches.insert(new_index, branches.pop(index))
            config['分支'] = {name: value for name, value in branches}
            with open('config.ini', 'w', encoding='utf-8') as configfile:
                config.write(configfile)
            return True
        return False

    def ini_to_excel(self, wb: Workbook):
        # 读取ini文件
        config = self.get_config()
        # 创建一个名为“设置”的工作表
        ws = wb.create_sheet("设置")
        # 写入ini文件内容到Excel
        row = 1
        for section in config.sections():
            ws.cell(row=row, column=1, value=f"[{section}]")
            row += 1
            for key, value in config.items(section):
                ws.cell(row=row, column=1, value=key)
                ws.cell(row=row, column=2, value=value)
                row += 1
            row += 1  # 在每个section后面加一个空行

    @staticmethod
    def excel_to_ini(wb: Workbook, ini_path: str = 'config.ini'):
        try:
            # 创建configparser对象
            config = configparser.ConfigParser()
            # 如果INI文件存在，则读取现有文件
            if os.path.exists(ini_path):
                config.read(ini_path, encoding='utf-8')
            ws = wb['设置']
            current_section = None
            # 读取Excel文件中的设置
            for row in ws.iter_rows(values_only=True):
                if row[0] is None:
                    continue
                if row[0].startswith('[') and row[0].endswith(']'):
                    # 这是一个section
                    current_section = row[0][1:-1]
                    if not config.has_section(current_section):
                        config.add_section(current_section)
                elif current_section and row[0] and row[1]:
                    # 这是一个键值对
                    config.set(current_section, str(row[0]), str(row[1]))
            # 将内容写入INI文件
            with open(ini_path, 'w', encoding='utf-8') as configfile:
                config.write(configfile)
        except Exception as e:
            print(f"设置写入失败: {e}")

    def set_current_branch(self, branch_name: str):
        """设置当前分支"""
        try:
            config = self.get_config()
            config.set("Config", "当前分支", branch_name)
            with open("config.ini", "w", encoding="utf-8") as f:
                config.write(f)
        except Exception as e:
            print("设置当前分支失败！", e)

    def get_current_branch(self) -> str:
        """获取当前分支"""
        try:
            config = self.get_config()
            return config["Config"]["当前分支"]
        except Exception as e:
            print("获取当前分支失败！", e)
            return "主流程"

    def system_prompt_tone(self, judge: str):
        """系统提示音
        :param judge: 判断类型（线程结束、全局快捷键、执行异常）"""
        try:
            is_tone = eval(self.get_setting_data_from_ini('Config', '系统提示音'))
            if judge == '线程结束' and is_tone:
                for i_ in range(3):
                    winsound.Beep(500, 300)
            elif judge == '全局快捷键' and is_tone:
                winsound.Beep(500, 300)
            elif judge == '执行异常' and is_tone:
                winsound.Beep(1000, 1000)
        except Exception as e:
            print('系统提示音错误！', e)

    def show_normal_window_with_specified_title(self, title):
        """将指定标题的窗口正常显示"""

        def get_window_titles(hwnd, titles):
            titles[hwnd] = win32gui.GetWindowText(hwnd)

        if eval(self.get_setting_data_from_ini('Config', '任务完成后显示主窗口')):
            hwnd_title = {}
            win32gui.EnumWindows(get_window_titles, hwnd_title)

            for h, t in hwnd_title.items():
                if t == title:
                    try:
                        time.sleep(0.5)
                        win32gui.ShowWindow(h, win32con.SW_SHOWNORMAL)  # 正常显示窗口
                    except Exception as e:
                        print(f"主窗口显示出现错误: {e}")
                    break


if __name__ == "__main__":
    pass
