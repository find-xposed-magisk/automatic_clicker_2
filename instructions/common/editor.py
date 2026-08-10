"""独立 .ui 参数窗口的通用绑定与校验。"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
import re

from PySide6.QtCore import QPoint, QRect, Qt, Signal
from PySide6.QtGui import QCursor
from PySide6.QtWidgets import (
    QApplication,
    QCheckBox,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QDoubleSpinBox,
    QFileDialog,
    QInputDialog,
    QLineEdit,
    QPlainTextEdit,
    QPushButton,
    QMessageBox,
    QRubberBand,
    QSpinBox,
    QWidget,
)

from instructions.base import InstructionEditorInterface
from instructions.models import InstructionDraft


@dataclass(frozen=True, slots=True)
class FieldSpec:
    key: str
    label: str
    kind: str = "text"
    default: Any = ""
    choices: tuple[str, ...] = ()
    required: bool = False
    minimum: float = -1_000_000
    maximum: float = 1_000_000


class _RegionSelectionDialog(QDialog):
    """鼠标框选屏幕区域的轻量覆盖层。"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowFlags(
            Qt.WindowType.FramelessWindowHint
            | Qt.WindowType.Tool
            | Qt.WindowType.WindowStaysOnTopHint
        )
        screen_ = QApplication.primaryScreen()
        if screen_ is not None:
            self.setGeometry(screen_.virtualGeometry())
        self.setWindowOpacity(0.28)
        self.setCursor(Qt.CursorShape.CrossCursor)
        self.setStyleSheet("background: black;")
        self._origin = QPoint()
        self._selection = QRect()
        self._dragging = False
        self._rubber_band = QRubberBand(QRubberBand.Shape.Rectangle, self)

    def mousePressEvent(self, event_) -> None:
        if event_.button() == Qt.MouseButton.LeftButton:
            self._dragging = True
            self._origin = event_.position().toPoint()
            self._rubber_band.setGeometry(QRect(self._origin, self._origin))
            self._rubber_band.show()

    def mouseMoveEvent(self, event_) -> None:
        if self._dragging:
            self._rubber_band.setGeometry(
                QRect(self._origin, event_.position().toPoint()).normalized()
            )

    def mouseReleaseEvent(self, event_) -> None:
        if event_.button() != Qt.MouseButton.LeftButton:
            return
        self._dragging = False
        local_rect_ = self._rubber_band.geometry().normalized()
        self._rubber_band.hide()
        if local_rect_.width() < 2 or local_rect_.height() < 2:
            self.reject()
            return
        top_left_ = self.mapToGlobal(local_rect_.topLeft())
        self._selection = QRect(top_left_, local_rect_.size())
        self.accept()

    def selected_region(self) -> tuple[int, int, int, int] | None:
        if self.result() != QDialog.DialogCode.Accepted:
            return None
        return (
            self._selection.x(),
            self._selection.y(),
            self._selection.width(),
            self._selection.height(),
        )


class SchemaInstructionEditor(QDialog, InstructionEditorInterface):
    """只绑定各指令 .ui 中已存在的控件，不创建窗口或参数控件。"""

    test_requested = Signal(object)
    auxiliary_requested = Signal(str, object)
    TYPE_ID = ""
    DISPLAY_NAME = ""
    FIELDS: tuple[FieldSpec, ...] = ()
    UI_CLASS = None

    def __init__(self, parent=None, draft=None, context=None):
        super().__init__(parent)
        if self.UI_CLASS is None:
            raise TypeError(f"{type(self).__name__} 未绑定独立 Ui 类")
        self.context = context
        self.ui = self.UI_CLASS()
        self.ui.setupUi(self)
        self.setWindowTitle(self.DISPLAY_NAME or self.TYPE_ID)
        self._controls = self._collect_parameter_controls()
        self._verify_ui_contract()
        self._bind_common_controls()
        self._connect_auxiliary_buttons()
        self._connect_buttons()
        if draft is not None:
            self.load_draft(draft)

    def _collect_parameter_controls(self) -> dict[str, QWidget]:
        controls_: dict[str, QWidget] = {}
        for index_, field_ in enumerate(self.FIELDS):
            control_name_ = f"parameter_{index_}"
            control_ = getattr(self.ui, control_name_, None)
            if not isinstance(control_, QWidget):
                raise RuntimeError(
                    f"{self.TYPE_ID}.ui 缺少参数控件：{control_name_} ({field_.key})"
                )
            controls_[field_.key] = control_
        return controls_

    def _verify_ui_contract(self) -> None:
        declared_ = {field_.key for field_ in self.FIELDS}
        actual_ = set(self._controls)
        if declared_ != actual_:
            missing_ = sorted(declared_ - actual_)
            extra_ = sorted(actual_ - declared_)
            raise RuntimeError(
                f"{self.TYPE_ID}.ui 参数控件与字段声明不一致；"
                f"缺少={missing_}，多余={extra_}"
            )

    def _bind_common_controls(self) -> None:
        required_names_ = (
            "repeatSpinBox",
            "errorPolicyComboBox",
            "noteEdit",
            "testButton",
            "buttonBox",
        )
        missing_ = [name_ for name_ in required_names_ if not hasattr(self.ui, name_)]
        if missing_:
            raise RuntimeError(f"{self.TYPE_ID}.ui 缺少通用控件：{missing_}")
        self.repeat_count = self.ui.repeatSpinBox
        self.error_policy = self.ui.errorPolicyComboBox
        self.note = self.ui.noteEdit
        self.test_button = self.ui.testButton
        self.button_box = self.ui.buttonBox

    def _connect_buttons(self) -> None:
        self.button_box.accepted.connect(self._accept_if_valid)
        self.button_box.rejected.connect(self.reject)
        self.test_button.clicked.connect(self._test_if_valid)

    def _connect_auxiliary_buttons(self) -> None:
        for index_, field_ in enumerate(self.FIELDS):
            button_ = getattr(self.ui, f"auxiliary_{index_}", None)
            if isinstance(button_, QPushButton):
                button_.clicked.connect(
                    lambda checked_=False, key__=field_.key: self._run_auxiliary(key__)
                )

    def _run_auxiliary(self, key_: str) -> None:
        control_ = self._controls[key_]
        current_ = self._control_value(control_)
        service_ = self.context.service("编辑器辅助") if self.context is not None else None
        if service_ is not None:
            result_ = service_(
                type_id=self.TYPE_ID,
                parameter_key=key_,
                current_value=current_,
                parent=self,
            )
            if result_ is not None:
                self._set_control_value(control_, result_)
                return

        if key_ in {"图像路径", "工作簿", "文件路径", "保存路径"}:
            self._browse_path(key_, control_)
            return
        if key_ == "代码":
            file_path_, _ = QFileDialog.getOpenFileName(
                self, "载入 Python 代码", "", "Python (*.py);;文本文件 (*.txt);;所有文件 (*)"
            )
            if file_path_:
                self._set_control_value(
                    control_, Path(file_path_).read_text(encoding="utf-8")
                )
            return
        if key_ == "区域":
            selector_ = _RegionSelectionDialog(self)
            selector_.exec()
            region_ = selector_.selected_region()
            if region_ is not None:
                self._set_control_value(control_, ",".join(str(value_) for value_ in region_))
            return
        if key_ == "单元格":
            cell_, accepted_ = QInputDialog.getText(
                self, "选择单元格", "请输入 Excel 单元格，例如 A1：", text=str(current_ or "A1")
            )
            if accepted_:
                if re.fullmatch(r"[A-Za-z]{1,3}[1-9][0-9]*", cell_.strip()):
                    self._set_control_value(control_, cell_.strip().upper())
                else:
                    QMessageBox.warning(self, "单元格无效", "请输入类似 A1、BC12 的单元格地址。")
            return
        if key_ == "变量":
            variables_ = sorted(str(name_) for name_ in getattr(self.context, "variables", {}))
            if not variables_:
                QMessageBox.information(self, "变量", "当前变量池为空。")
                return
            variable_, accepted_ = QInputDialog.getItem(
                self, "选择变量", "变量名称：", variables_, editable=False
            )
            if accepted_:
                self._set_control_value(control_, variable_)
            return
        if key_ in {"坐标", "开始位置", "结束位置", "点击位置"}:
            position_ = QCursor.pos()
            separator_ = "-" if self.TYPE_ID == "坐标点击" and key_ == "坐标" else ","
            self._set_control_value(control_, f"{position_.x()}{separator_}{position_.y()}")
            return
        self.auxiliary_requested.emit(key_, control_)

    def _browse_path(self, key_: str, control_: QWidget) -> None:
        if key_ == "保存路径":
            file_path_, _ = QFileDialog.getSaveFileName(
                self, "选择保存路径", "", "PNG 图像 (*.png);;所有文件 (*)"
            )
            paths_ = [file_path_] if file_path_ else []
        elif key_ == "工作簿":
            file_path_, _ = QFileDialog.getOpenFileName(
                self, "选择工作簿", "", "Excel 工作簿 (*.xlsx *.xlsm);;所有文件 (*)"
            )
            paths_ = [file_path_] if file_path_ else []
        elif isinstance(control_, QPlainTextEdit):
            paths_, _ = QFileDialog.getOpenFileNames(
                self, "选择文件", "", "图像 (*.png *.jpg *.jpeg *.bmp);;所有文件 (*)"
            )
        else:
            file_path_, _ = QFileDialog.getOpenFileName(self, "选择文件")
            paths_ = [file_path_] if file_path_ else []
        if paths_:
            value_ = "\n".join(paths_) if isinstance(control_, QPlainTextEdit) else paths_[0]
            self._set_control_value(control_, value_)
            if key_ == "工作簿" and "工作表" in self._controls:
                try:
                    from openpyxl import load_workbook
                    workbook_ = load_workbook(paths_[0], read_only=True, data_only=True)
                    try:
                        if workbook_.sheetnames:
                            self._set_control_value(
                                self._controls["工作表"], workbook_.sheetnames[0]
                            )
                    finally:
                        workbook_.close()
                except Exception as error_:
                    QMessageBox.warning(self, "工作簿读取失败", str(error_))

    def _accept_if_valid(self) -> None:
        try:
            self.get_draft()
        except (TypeError, ValueError) as error_:
            QMessageBox.warning(self, "参数不完整", str(error_))
            return
        self.accept()

    def _test_if_valid(self) -> None:
        try:
            draft_ = self.get_draft()
        except (TypeError, ValueError) as error_:
            QMessageBox.warning(self, "参数不完整", str(error_))
            return
        self.test_requested.emit(draft_)

    @staticmethod
    def _control_value(control_: QWidget):
        if isinstance(control_, QCheckBox):
            return control_.isChecked()
        if isinstance(control_, (QSpinBox, QDoubleSpinBox)):
            return control_.value()
        if isinstance(control_, QComboBox):
            return control_.currentText()
        if isinstance(control_, QPlainTextEdit):
            return control_.toPlainText()
        if isinstance(control_, QLineEdit):
            return control_.text()
        raise TypeError(f"不支持的参数控件：{type(control_).__name__}")

    @staticmethod
    def _set_control_value(control_: QWidget, value_: Any) -> None:
        if isinstance(control_, QCheckBox):
            control_.setChecked(bool(value_))
        elif isinstance(control_, (QSpinBox, QDoubleSpinBox)):
            control_.setValue(value_)
        elif isinstance(control_, QComboBox):
            text_ = str(value_)
            if control_.findText(text_) < 0:
                control_.addItem(text_)
            control_.setCurrentText(text_)
        elif isinstance(control_, QPlainTextEdit):
            control_.setPlainText(str(value_ or ""))
        elif isinstance(control_, QLineEdit):
            control_.setText(str(value_ or ""))
        else:
            raise TypeError(f"不支持的参数控件：{type(control_).__name__}")

    def get_draft(self) -> InstructionDraft:
        parameters_ = {}
        for field_ in self.FIELDS:
            value_ = self._control_value(self._controls[field_.key])
            if field_.required and (value_ is None or str(value_).strip() == ""):
                raise ValueError(f"{field_.label}不能为空")
            parameters_[field_.key] = value_
        self._validate_parameters(parameters_)
        return InstructionDraft(
            type_id=self.TYPE_ID,
            parameters=parameters_,
            repeat_count=self.repeat_count.value(),
            error_policy=self.error_policy.currentText(),
            note=self.note.text(),
        )

    def _validate_parameters(self, parameters_: dict[str, Any]) -> None:
        """在生成 Draft 前验证跨控件约束和文本格式。"""
        for field_ in self.FIELDS:
            value_ = parameters_[field_.key]
            if field_.kind == "choice" and field_.choices and value_ not in field_.choices:
                raise ValueError(f"{field_.label}不是有效选项")

        for key_ in ("坐标", "开始位置", "结束位置", "点击位置"):
            value_ = parameters_.get(key_)
            if value_ in (None, ""):
                continue
            if key_ == "点击位置" and str(value_).replace(" ", "").strip("()") == "随机,随机":
                continue
            try:
                from instructions.common.actions import point

                point(value_)
            except (TypeError, ValueError) as error_:
                raise ValueError(f"{key_}必须为 x,y 或 x-y") from error_

        region_ = parameters_.get("区域")
        if region_ not in (None, ""):
            try:
                from instructions.common.actions import region

                parsed_region_ = region(region_)
            except (TypeError, ValueError) as error_:
                raise ValueError("区域必须为 x,y,width,height") from error_
            if parsed_region_ is not None and (parsed_region_[2] <= 0 or parsed_region_[3] <= 0):
                raise ValueError("区域宽度和高度必须大于 0")

        cell_ = parameters_.get("单元格")
        if cell_ not in (None, "") and not re.fullmatch(
            r"[A-Za-z]{1,3}[1-9][0-9]*", str(cell_).strip()
        ):
            raise ValueError("单元格必须为类似 A1、BC12 的地址")

        target_time_ = parameters_.get("时间")
        if target_time_ not in (None, ""):
            try:
                datetime.strptime(str(target_time_), "%H:%M:%S")
            except ValueError as error_:
                raise ValueError("目标时间必须为 HH:MM:SS") from error_

        if parameters_.get("类型") == "随机等待":
            factor_ = {"毫秒": 0.001, "秒": 1.0, "分钟": 60.0}
            minimum_ = float(parameters_.get("最小", 0)) * factor_[
                str(parameters_.get("最小单位", "秒"))
            ]
            maximum_ = float(parameters_.get("最大", 0)) * factor_[
                str(parameters_.get("最大单位", "秒"))
            ]
            if minimum_ > maximum_:
                raise ValueError("随机等待最小值不能大于最大值")

        if parameters_.get("类型") == "随机滚轮滑动":
            if int(parameters_.get("最小距离", 0)) > int(parameters_.get("最大距离", 0)):
                raise ValueError("随机滚轮最小距离不能大于最大距离")

    def load_draft(self, draft) -> None:
        draft_ = draft if isinstance(draft, InstructionDraft) else InstructionDraft.from_mapping(draft)
        if draft_.type_id != self.TYPE_ID:
            raise ValueError(f"不能用 {draft_.type_id} 初始化 {self.TYPE_ID} 编辑器")
        for key_, value_ in draft_.parameters.items():
            control_ = self._controls.get(key_)
            if control_ is not None:
                self._set_control_value(control_, value_)
        self.repeat_count.setValue(draft_.repeat_count)
        self.error_policy.setCurrentText(draft_.error_policy)
        self.note.setText(draft_.note)
