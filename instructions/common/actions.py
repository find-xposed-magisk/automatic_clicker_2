"""可复用的运行期工具；具体指令在各自模块选择和调用这些工具。"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Callable
import contextlib
import io
import os
import random
import subprocess
import time

from instructions.models import ExecutionContext


def parameter(parameters: dict, *names: str, default=None):
    for name_ in names:
        if name_ in parameters:
            return parameters[name_]
    return default


def service_or(context: ExecutionContext, name: str, fallback: Callable, **parameters):
    service_ = context.service(name)
    if service_ is not None:
        return service_(**parameters)
    return fallback(**parameters)


def delegated(context: ExecutionContext, name: str, command):
    """调用宿主注入的完整指令实现；返回 (是否已处理, 返回值)。"""
    service_ = context.service(name)
    if service_ is None:
        return False, None
    return True, service_(context=context, command=command)


def pyautogui_module():
    import pyautogui
    return pyautogui


def point(value_: Any) -> tuple[int, int]:
    if isinstance(value_, (list, tuple)) and len(value_) >= 2:
        return int(value_[0]), int(value_[1])
    if isinstance(value_, dict):
        return int(value_.get("x", 0)), int(value_.get("y", 0))
    text_ = str(value_ or "0,0").strip().strip("()[]")
    text_ = text_.replace("，", ",")
    if "," in text_:
        x_, y_ = text_.split(",", maxsplit=1)
    elif "-" in text_.lstrip("-"):
        separator_index_ = text_.lstrip("-").index("-") + (1 if text_.startswith("-") else 0)
        x_, y_ = text_[:separator_index_], text_[separator_index_ + 1:]
    else:
        raise ValueError("坐标必须为 x,y 或 x-y")
    return int(float(x_.strip())), int(float(y_.strip()))


def region(value_: Any):
    if value_ in (None, "", [], ()):
        return None
    if isinstance(value_, (list, tuple)) and len(value_) == 4:
        return tuple(int(v_) for v_ in value_)
    text_ = str(value_).strip().strip("()[]")
    values_ = [int(float(item_)) for item_ in text_.replace("，", ",").split(",")]
    if len(values_) != 4:
        raise ValueError("区域必须为 x,y,width,height")
    return None if all(value_ == 0 for value_ in values_) else tuple(values_)


def resolve_image_path(parameters: dict, context: ExecutionContext | None = None) -> str:
    image_ = parameter(parameters, "图像路径", "图像", "图片", default="")
    if not image_:
        raise ValueError("未设置图像路径")
    image_path_ = substitute_variables(context, str(image_)) if context is not None else str(image_)
    if not os.path.isfile(image_path_) and context is not None:
        database_ = context.metadata.get("database")
        if database_ is not None:
            for folder_ in database_.extract_resource_folder_path():
                candidate_ = os.path.join(folder_, os.path.basename(image_path_))
                if os.path.isfile(candidate_):
                    image_path_ = candidate_
                    break
    return image_path_


def locate_image(
    parameters: dict,
    context: ExecutionContext | None = None,
    *,
    min_search_time: float = 0.0,
):
    image_path_ = resolve_image_path(parameters, context)
    confidence_ = float(parameter(parameters, "精度", default=0.8))
    if confidence_ > 1:
        confidence_ /= 100
    try:
        return pyautogui_module().locateCenterOnScreen(
            image_path_,
            confidence=confidence_,
            grayscale=bool(parameter(parameters, "灰度", default=False)),
            region=region(parameter(parameters, "区域", default=None)),
            minSearchTime=max(0.0, float(min_search_time)),
        )
    except (FileNotFoundError, OSError):
        return None
    except Exception as error_:
        if error_.__class__.__name__ == "ImageNotFoundException":
            return None
        raise


def image_error_timeout(parameters: dict) -> tuple[bool, float]:
    """返回（未找到时自动跳过、最长识别秒数）。"""
    policy_ = parameter(
        parameters,
        "图像等待",
        "图像超时",
        "查找超时",
        "异常",
        default="自动略过",
    )
    if policy_ in (None, "", "自动跳过", "自动略过"):
        return True, 1.0
    try:
        return False, max(0.0, float(policy_))
    except (TypeError, ValueError):
        return False, 0.0


def locate_image_with_policy(parameters: dict, context: ExecutionContext | None = None):
    skip_, timeout_ = image_error_timeout(parameters)
    return locate_image(parameters, context, min_search_time=timeout_), skip_


def image_random_offset(parameters: dict, context: ExecutionContext | None = None) -> tuple[int, int]:
    """在目标图像实际边界内生成相对于中心的随机位置。"""
    from PIL import Image

    image_path_ = resolve_image_path(parameters, context)
    with Image.open(image_path_) as image_:
        width_, height_ = image_.size
    return (
        random.randint(-(width_ // 2), width_ // 2),
        random.randint(-(height_ // 2), height_ // 2),
    )


def mouse_action(action_: str, x_: int, y_: int, count_: int | None = None, interval_: float = 0.0):
    gui_ = pyautogui_module()
    action_map_ = {
        "左键单击": ("left", 1), "左键双击": ("left", 2), "左键三击": ("left", 3),
        "右键单击": ("right", 1), "右键双击": ("right", 2), "中键单击": ("middle", 1),
        "仅移动鼠标": (None, 0),
    }
    is_standard_action_ = action_ in action_map_
    button_, default_count_ = action_map_.get(action_, ("left", 1))
    gui_.moveTo(x_, y_)
    if button_ is not None:
        click_count_ = default_count_ if is_standard_action_ or count_ is None else int(count_)
        gui_.click(x_, y_, clicks=max(1, click_count_), interval=float(interval_), button=button_)


def wait_seconds(seconds_: float) -> None:
    time.sleep(max(0.0, float(seconds_)))


def workbook_cell(path_: str, sheet_: str, cell_: str, data_only: bool = True):
    from openpyxl import load_workbook
    workbook_ = load_workbook(path_, data_only=data_only)
    try:
        worksheet_ = workbook_[sheet_] if sheet_ else workbook_.active
        return workbook_, worksheet_, worksheet_[cell_]
    except Exception:
        workbook_.close()
        raise


def store_variable(context: ExecutionContext, parameters: dict, value_: Any):
    name_ = str(parameter(parameters, "变量", "变量名称", default=""))
    return context.set_variable(name_, value_)


def substitute_variables(context: ExecutionContext, text_: str) -> str:
    result_ = str(text_)
    for name_, value_ in context.variables.items():
        result_ = result_.replace(f"{{{{{name_}}}}}", str(value_))
        result_ = result_.replace(f"☾{name_}☽", str(value_))
    return result_


def substitute_python_variables(context: ExecutionContext, code_: str) -> str:
    """按旧协议把 ☾变量☽ 替换为可安全嵌入 Python 的字面量。"""
    result_ = str(code_)
    for name_, value_ in context.variables.items():
        result_ = result_.replace(f"☾{name_}☽", repr(value_))
    return result_


def run_python_code(
    context: ExecutionContext,
    code_: str,
    result_name_: str = "",
    variable_name_: str = "",
):
    namespace_ = {"context": context, "variables": context.variables}
    output_ = io.StringIO()
    with contextlib.redirect_stdout(output_):
        exec(
            compile(substitute_python_variables(context, code_), "<clicker-instruction>", "exec"),
            namespace_,
            namespace_,
        )
    if result_name_:
        value_ = namespace_.get(result_name_)
    else:
        value_ = namespace_.get("result", output_.getvalue().rstrip())
    if variable_name_:
        context.set_variable(variable_name_, value_)
    return value_


def run_process(command_: Any, *, shell_: bool = False):
    return subprocess.run(command_, shell=shell_, check=True, capture_output=True, text=True)


def current_time(format_: str):
    aliases_ = {
        "年-月-日 小时:分钟:秒": "%Y-%m-%d %H:%M:%S",
        "年/月/日 小时:分钟:秒": "%Y/%m/%d %H:%M:%S",
        "月/日/年 小时:分钟:秒": "%m/%d/%Y %H:%M:%S",
        "日-月-年 小时:分钟:秒": "%d-%m-%Y %H:%M:%S",
        "年-月-日 时:分:秒": "%Y-%m-%d %H:%M:%S",
        "年-月-日": "%Y-%m-%d",
        "月/日/年": "%m/%d/%Y",
        "日-月-年": "%d-%m-%Y",
        "年-月": "%Y-%m",
        "月/年": "%m/%Y",
        "年": "%Y",
        "时:分:秒": "%H:%M:%S",
        "时间戳": "%s",
    }
    format__ = aliases_.get(format_, format_ or "%Y-%m-%d %H:%M:%S")
    if format__ == "%s":
        return str(int(time.time()))
    return datetime.now().strftime(format__)


def resolve_increment(cell_: str, increment_: bool, iteration_: int) -> str:
    if not increment_ or iteration_ <= 1:
        return cell_
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter
    column_, row_ = coordinate_from_string(cell_.upper())
    return f"{get_column_letter(column_index_from_string(column_))}{row_ + iteration_ - 1}"


def random_offset(enabled_: bool, radius_: int = 100) -> tuple[int, int]:
    if not enabled_:
        return 0, 0
    return random.randint(-radius_, radius_), random.randint(-radius_, radius_)


def ensure_parent(path_: str) -> None:
    Path(path_).expanduser().resolve().parent.mkdir(parents=True, exist_ok=True)
